from openpyxl import load_workbook
class laptop:
    symbol = ""
    name = ""
    count = 0
    description = ""
    keyboard = ""
    netto = 0
    price_with_delivery = 0

    def __init__(self, symbol = "", name="", count=0, description="", 
                        keyboard="", netto=0, price_with_delivery=0):
        self.symbol = symbol
        self.name = name
        self.count = count
        self.description = description
        self.keyboard = keyboard
        self.netto = netto
        self.price_with_delivery = price_with_delivery

laptops = laptop()

def Clear():
    wb = load_workbook('notebooks_vasyl.xlsx')
    sheet = wb.active
    
    # Clear
    maxColumn = int(sheet.max_column + 1)
    maxRow = int(sheet.max_row + 1)
    for i in range(1, maxColumn):
        for j in range(2, maxRow):
            sheet.cell(column=i, row=j).value = None 
    #
    wb.save("notebooks_vasyl.xlsx") 

def FromNotebooks():

    #load from notebooks
    wb = load_workbook('notebooks.xlsx')
    sheet = wb.active
    
    for l in range(2, int(sheet.max_row)+1):
            obj = laptop(symbol = sheet.cell(column = 1, row = l).value, 
                        name = sheet.cell(column = 2, row = l).value,
                        count = sheet.cell(column = 3, row = l).value,                  
                        netto = sheet.cell(column = 4, row = l).value,
                        description = sheet.cell(column = 5, row = l).value)                   
            laptops.append(obj) 
    
# clear file notebooks_vasyl.xlsx
Clear()
# FromNotebooks()

