from AGD import AGD
import os
from openpyxl import load_workbook

# exists file
def exists(path):
    try:
        os.stat(path)
    except OSError:
        print(f"File: {path} isn`t exists")
        return False
    # print(f"File: {path} is exists")
    return True
#convert fild PRICE in normal number

def FromBiss(agd):
    if exists('BISS/BISS.xlsx'):
        wb = load_workbook('BISS/BISS.xlsx')
        sheet = wb.active  
        start = 2
        if sheet.cell(column = 1, row = start).value == "Nazwa": start = 3

        # print("max row in BISS =" + str(sheet.max_row))
        for l in range(start, int(sheet.max_row)+1):
            tmp_price = 0
            tmp_count = 0
            tmp_rezervacion = 0
            if sheet.cell(column = 6, row = l).value != None: tmp_price = int(sheet.cell(column = 6, row = l).value)
            if sheet.cell(column = 4, row = l).value != None: tmp_count = int(sheet.cell(column = 4, row = l).value)
            if sheet.cell(column = 3, row = l).value != None: tmp_rezervacion = int(sheet.cell(column = 3, row = l).value)
            # print(f"try add {l}")
            

            obj = AGD(name = sheet.cell(column = 1, row = l).value, 
                        count = tmp_count,
                        rezervacion = tmp_rezervacion,
                        price = tmp_price,
                        sklad = 'biss')
            agd.append(obj)
    return agd
