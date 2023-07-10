from AGD import AGD
import os
from openpyxl import load_workbook

# exists file
def exists(path):
    try:
        os.stat(path)
    except OSError:
        print(f"File: {path} isn`t exists")
        return False
    # print(f"File: {path} is exists")
    return True

#convert fild PRICE in normal number
def PriceToBrutto(price):
    if price[0] == '.': return 0
    price = price.replace("  ", " ")
    price = price.split(' ')
    return price[0]

def FromMarpa(agd):
    if exists('MARPA/MARPA.xlsx'):
        wb = load_workbook('MARPA/MARPA.xlsx')
        sheet = wb.active  

        for l in range(2, int(sheet.max_row)+1):
            obj = AGD(name = sheet.cell(column = 1, row = l).value, 
                        count = sheet.cell(column = 2, row = l).value,
                        rezervacion = sheet.cell(column = 3, row = l).value,
                        description = sheet.cell(column = 4, row = l).value,
                        price = int(PriceToBrutto(sheet.cell(column = 5, row = l).value)),
                        country = sheet.cell(column = 6, row = l).value,
                        sklad = 'marpa')
            agd.append(obj)
    return agd
