from openpyxl import load_workbook
import csv

class laptop:
    symbol = ""
    name = ""
    count = 0
    description = ""
    keyboard = ""
    netto = 0
    price_with_delivery = 0

    def __init__(self, symbol = "", name="", count=0, description="", 
                        keyboard="", netto=0, price_with_delivery=0):
        self.symbol = symbol
        self.name = name
        self.count = count
        self.description = description
        self.keyboard = keyboard
        self.netto = netto
        self.price_with_delivery = price_with_delivery

    def printNotebooks(self):
        print(f"symbol= {self.symbol}\n\
                    name= {self.name}\n\
                    count= {self.count}\n\
                    keyboard = {self.keyboard}\n\
                    netto= {self.netto}\n\
                    price_with_delivery= {self.price_with_delivery}\n\
                    description = {self.description}")
    def GetSymbol(self): return self.symbol
    def GetName(self): return self.name
    def GetCount(self): return self.count
    def GetDescription(self): return self.description
    def GetKeyboard(self): return self.keyboard
    def GetNetto(self): return self.netto
    def GetPriceWithDelivery(self): return self.price_with_delivery

    # def SetSymbol(self, symbol): self.symbol = symbol
    # def SetName(self, name): self.name = name
    # def SetCount(self, count): self.count = count
    # def SetDescription(self, description): self.description = description
    # def SetKeyboard(self, keyboard): self.keyboard = keyboard
    # def SetNetto(self, netto): self.netto = netto
    # def SetPriceWithDelivery(self, percent): 
    #     self.price_with_delivery = round(int(self.netto*(percent/100+1))) 
# load keyboard from file localization.csv 
    def Keyboard(self):
        file = open("notebooks/localization.csv", 'r')
        key_lines = csv.reader(file)

        keyb = []
        for i in key_lines:
            keyb.append(",".join(i))
        file.close()

        #write keyboard in object of laptop 
        for i in keyb:
            if str(self.symbol[-3:]) == str(i[0:3]):
                self.keyboard = i[4:]
                break

laptops = []

def Clear():
    wb = load_workbook('price_notebooks_of_vasyl.xlsx')
    sheet = wb.active
    
    # Clear
    maxColumn = int(sheet.max_column + 1)
    maxRow = int(sheet.max_row + 1)
    print(maxRow)
    print(maxColumn)
    for i in range(1, maxColumn):
        for j in range(2, maxRow):
            sheet.cell(column=i, row=j).value = None
    #
    wb.save("price_notebooks_of_vasyl.xlsx") 



def FromNotebooks():

    #load from notebooks
    wb = load_workbook('notebooks/notebooks.xlsx')
    sheet = wb.active
    
    for l in range(2, int(sheet.max_row)+1):
            # obj = laptop()
            # laptop.SetSymbol(l,sheet.cell(column = 1, row = l).value)

            obj = laptop(symbol = sheet.cell(column = 1, row = l).value, 
                    name = sheet.cell(column = 2, row = l).value,
                    count = sheet.cell(column = 3, row = l).value,                  
                    netto = sheet.cell(column = 4, row = l).value,
                    description = sheet.cell(column = 6, row = l).value,
                    price_with_delivery = round(int(sheet.cell(column = 4, row = l).value*1.2))) #percent                 
            laptops.append(obj)

    print(f"Loaded {len(laptops)} laptops")
    # insert keyboard i list of object
    for i in laptops:
        laptop.Keyboard(i)
    
        wb = load_workbook('price_notebooks_of_vasyl.xlsx')
    sheet = wb.active

    #write price from obj
    count = 2
    for i in laptops:
        if i.symbol[-3:] == "A2N": continue # if keyboard==A2N - Arabic than do not write
        sheet.cell(row=count, column=1).value = count-1
        sheet.cell(row=count, column=2).value = i.GetSymbol()
        sheet.cell(row=count, column=3).value = i.GetKeyboard()
        sheet.cell(row=count, column=4).value = i.GetName()
        sheet.cell(row=count, column=5).value = i.GetDescription()
        sheet.cell(row=count, column=6).value = i.GetCount()
        sheet.cell(row=count, column=8).value = i.GetPriceWithDelivery()
        formula_uah=f"=H{count}*$J$1"
        sheet.cell(row=count, column=9).value = formula_uah
        count += 1
    #
    wb.save("price_notebooks_of_vasyl.xlsx") 
    print(f"Write to price {count-2} laptops")
    
Clear()
FromNotebooks()

