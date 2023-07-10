from AGD import AGD
import os
from openpyxl import load_workbook

# exists file
def exists(path):
    try:
        os.stat(path)
    except OSError:
        print(f"File: {path} isn`t exists")
        return False
    # print(f"File: {path} is exists")
    return True
#convert fild PRICE in normal number

def FromNexa(agd):
    if exists('NEXA/NEXA.xlsx'):
        wb = load_workbook('NEXA/NEXA.xlsx')
        sheet = wb.active  

        for l in range(2, int(sheet.max_row)):
            # print(f"try add {l}")
            obj = AGD(name = sheet.cell(column = 2, row = l).value, 
                        count = sheet.cell(column = 3, row = l).value,
                        # rezervacion = sheet.cell(column = 3, row = l).value,
                        # description = sheet.cell(column = 5, row = l).value,
                        price = sheet.cell(column = 4, row = l).value,
                        country = sheet.cell(column = 5, row = l).value,
                        sklad = 'nexa')
            agd.append(obj)
    return agd

def PriceToBrutto(price):
    price = price.replace("  ", " ")
    price = price.split(' ')
    return price[0]

def FromNexaZakaz(agd):
    if exists('NEXA/NEXA-ZAKAZ.xlsx'):
        wb = load_workbook('NEXA/NEXA-ZAKAZ.xlsx')
        sheet = wb.active  

        for l in range(2, int(sheet.max_row)):
            # print(f"try add {l}")
            name_temp = str(sheet.cell(column = 1, row = l).value)
            if len(name_temp) > 6:
                obj = AGD(name = name_temp, 
                        count = sheet.cell(column = 3, row = l).value,
                        price = sheet.cell(column = 4, row = l).value,
                        sklad = 'nexa Під замовлення')
                agd.append(obj)
    return agd
