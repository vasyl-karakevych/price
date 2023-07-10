from itertools import count
from MARPA.marpa import FromMarpa, exists
from NEXA.nexa import FromNexa, FromNexaZakaz, exists
from BISS.biss import FromBiss, exists
from AGD import AGD
from openpyxl import load_workbook
# from notebooks.notebooks import WriteToNotebooks

agd = []
delivery = 1.17
NORMA = 2200
freezer = 550
pralka = 450

#print object to console
def PrintAGD(obj):
    print(f"name= {obj.GetName()}\n\
            type= {obj.type}\n\
            count= {obj.count}\n\
            rezervacion= {obj.rezervacion}\n\
            description= {obj.description}\n\
            country= {obj.country}\n\
            price= {obj.GetPrice()}\n\
            price_with_delivery= {obj.GetPriceWD()}\n")

# count and add delivery to price
def PriceDelivery():

    for obj in agd:
        if   obj.name.find("CHŁODZ") >= 0 or obj.name.find("Chłodziarka") >=0 or obj.name.find("LODÓWKA") >=0 : obj.type = "Холодильник"
        elif obj.name.find("PRALKA") >= 0 or obj.name.find("Pralka") >= 0: obj.type = "Пралка"
        elif obj.name.find("PRALKO_SUSZ") >= 0: obj.type = "Пралко-сушарка"
        elif obj.name.find("TV ") >= 0: obj.type = "Телевізор"
    
    # створюємо словник із значеннями країн походження товару
    country_mapping = {
    "WŁOCHY": "Китай", "Włochy": "Китай", "CHINY": "Китай", "Chiny": "Китай",
    "SŁOWENIA": "Словенія", "Słowenia": "Словенія",
    "POLSKA": "Польща", "POLAND": "Польща", "Polska": "Польща",
    "RUMUNIA": "Румунія",
    "TURCJA": "Туреччина", "Turcja": "Туреччина",
    "SERBIA": "Сербія",
    "NIEMCY": "Німеччина", "Niemcy": "Німеччина",
    "INDONEZJA": "Індонезія",
    "UK": "Великобританія", "Wielka Brytania": "Великобританія",
    "EU": "Європа",
    "HISZPANIA": "Іспанія", "Hiszpania": "Іспанія",
    "FRANCJA": "Франція",
    "Grecja": "Греція",
    "Malesia": "Малайзія", "MALEZJA": "Малайзія",
    "Węgry": "Угорщина",
    "KOREA": "Корея",
    "SŁOWACJA": "Словаччина",
    "Wietnam": "В'єтнам"
    }    
 
    for obj in agd:
        # Пілставляємо значення із словника
        obj.country = country_mapping.get(obj.country, "-")
        
        # В прайсах на Marpi телевізори вказують у нетто. Відповідно нижче вираз міняє нетто на брутто ціну телевізорів.
        if obj.type == "Телевізор" and obj.sklad == "marpa": obj.price *= 1.23

        brutto = obj.price
        netto = brutto/1.23
        percent =  netto * (delivery-1)


        if brutto > 0:
            if obj.type == "Холодильник": 
                if brutto > 1000 and brutto < NORMA:
                    obj.price_with_delivery = round(netto + freezer)
                elif brutto > NORMA:
                    obj.price_with_delivery = round(netto + freezer + percent)
            elif obj.type == "Пралка" or obj.type == "Пралко-сушарка":
                if brutto < NORMA: 
                    obj.price_with_delivery = round(netto + pralka)
                elif brutto > NORMA:              
                    obj.price_with_delivery = round(netto + pralka + percent)
            else: obj.price_with_delivery = round(netto * delivery)

def WriteToPrice():
    wb = load_workbook('price_vasyl.xlsx')
    sheet = wb.active
    
    # Clear
    maxColumn = int(sheet.max_column + 1)
    maxRow = int(sheet.max_row + 1)
    for i in range(1, maxColumn):
        for j in range(2, maxRow):
            sheet.cell(column=i, row=j).value = None 

    counter = 1
    for obj in agd:
        if (obj.count >0):
            sheet.cell(column=1, row=counter+1).value = counter
            sheet.cell(column=2, row=counter+1).value = obj.GetName()
            sheet.cell(column=3, row=counter+1).value = obj.GetDescription()
            sheet.cell(column=4, row=counter+1).value = obj.GetCount()
            sheet.cell(column=5, row=counter+1).value = obj.GetRezervacion()
            sheet.cell(column=6, row=counter+1).value = obj.GetPrice()
            sheet.cell(column=7, row=counter+1).value = obj.price_with_delivery
            sheet.cell(column=8, row=counter+1).value = f"=G{counter+1}*$L$1"
            sheet.cell(column=9, row=counter+1).value = obj.GetSklad()
            sheet.cell(column=10, row=counter+1).value = obj.GetType()
            sheet.cell(column=11, row=counter+1).value = obj.GetCountry()
            counter += 1
    print(f"Write to price: {counter-1} objects")
    wb.save("price_vasyl.xlsx") 


# TotalCount inprice
max_len_price = len(agd)

# Load from MARPA
FromMarpa(agd)
len_marpa = len(agd) - max_len_price
print(f"Load from MARPA: {len_marpa} records")
max_len_price = len_marpa
# PrintAGD(agd[0])

# Load from NEXA
FromNexa(agd)
len_nexa = len(agd) - max_len_price
print(f"Load from Nexa: {len_nexa} records")
max_len_price += len_nexa

# Load from BISS
FromBiss(agd)
len_biss = len(agd) - max_len_price
print(f"Load from Biss: {len_biss} records")
max_len_price += len_biss

# Load from NEXA Zakaz
FromNexaZakaz(agd)
len_nexa_zakaz = len(agd) - max_len_price
print(f"Load from Nexa: {len_nexa_zakaz} records")
max_len_price += len_nexa_zakaz

print(f"Loaded: {max_len_price} records")

PriceDelivery()
# PrintAGD(agd[54])
WriteToPrice()

# LAPTOPS
# laptops = AGD()
# WriteToNotebooks(laptops)