from AGD import AGD
import os
from openpyxl import load_workbook

def exists(path):
    try:
        os.stat(path)
    except OSError:
        # print(f"File: {path} isn`t exists")
        return False
    print(f"File: {path} is exists")
    return True

#convert fild PRICE in normal number
def PriceToBrutto(price):
    price = price.split(' ')
    if len(price) > 1:
        if price[1][0] == 'B': price = int(price[0])
        elif price[1][0] == 'N': price = round(int(price[0])*1.23)
        else: price = 0
    else: price = 0
    return price

def FromMarpa(agd):
    if exists('MARPA.xlsx'):
        wb = load_workbook('MARPA.xlsx')
        sheet = wb.active  

        obj = AGD(name = sheet.cell(column = 1, row = 2).value, 
                    count = sheet.cell(column = 2, row = 2).value,
                    rezervacion = sheet.cell(column = 3, row = 2).value,
                    description = sheet.cell(column = 4, row = 2).value,
                    price = PriceToBrutto(sheet.cell(column = 5, row = 2).value),
                    country = sheet.cell(column = 6, row = 2).value)

    agd.append(obj)
    return agd
